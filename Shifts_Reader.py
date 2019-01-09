from openpyxl import load_workbook

def ParseReq(req):
    return req.split(",")

def readEmployeeReq(fileName, emp):
    """this function read all the requests of the employee and set it in current property
    input: DB FILE NAME , employee object"""
    wb = load_workbook('Requests/' + fileName + "/" +  emp.name +'.xlsx')

    ws = wb.active
    priorities = []
    requestList = []

    tupleList = []
    for row in ws.iter_cols(min_row=2,min_col=3,max_col=4,  max_row=8):     #read from reaquests file
       for cell in row:
           if (cell.col_idx == 3):
               priorities.append(ParseReq(cell.value))
           if (cell.col_idx  == 4):
               requestList.append(cell.value)

    workerTuple = []
    for i in range (0,7):
        workerTuple.append((priorities[i],requestList[i]))          #set in tuple



    emp.sft_request = workerTuple               #set in employee property